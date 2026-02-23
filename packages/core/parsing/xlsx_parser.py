"""XLSX parsing functions.

Extracts data from Excel (.xlsx) files using openpyxl.
Returns data in the same row format as csv_parser.parse_csv_content().
"""
import io
from typing import Any, Dict, List

import openpyxl

from packages.core.parsing.csv_parser import KNOWN_HEADERS


def parse_xlsx_content(
    content: bytes,
    sheet_name: str | None = None,
) -> List[Dict[str, Any]]:
    """Parse XLSX content into list of dictionaries.

    Uses openpyxl to read the workbook.  If *sheet_name* is ``None`` the
    first (active) sheet is used.

    Returns the same format as ``parse_csv_content()``::

        [{"row_number": int, "cells": List[str], "headers": List[str]}]

    Args:
        content: Raw XLSX bytes.
        sheet_name: Optional worksheet name to read.

    Returns:
        List of row dicts.

    Raises:
        ValueError: If the file cannot be parsed or contains no data.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as e:
        raise ValueError(f"Failed to open XLSX: {e}")

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(
                f"Sheet '{sheet_name}' not found. "
                f"Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    if ws is None:
        wb.close()
        raise ValueError("Workbook has no active sheet")

    # Read all rows into list of string lists
    all_rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append([_clean_cell(c) for c in row])

    wb.close()

    if not all_rows:
        raise ValueError("XLSX file is empty")

    # Detect header row (reuse CSV heuristic)
    header_idx = _detect_header_row(all_rows)
    headers = all_rows[header_idx]

    # Build output rows
    rows: List[Dict[str, Any]] = []
    for row_idx, raw_row in enumerate(
        all_rows[header_idx + 1:], start=header_idx + 2
    ):
        # Skip empty rows
        if not any(c.strip() for c in raw_row):
            continue

        # Pad/trim cells to match header count
        cells = raw_row[:len(headers)]
        while len(cells) < len(headers):
            cells.append("")

        rows.append({
            "row_number": row_idx,
            "cells": cells,
            "headers": headers,
        })

    if not rows:
        raise ValueError("XLSX file is empty or contains only headers")

    return rows


def _detect_header_row(all_rows: List[List[str]], max_rows: int = 20) -> int:
    """Find the row index that most likely contains headers.

    Same logic as csv_parser.detect_header_row but works on pre-parsed rows.
    """
    best_idx = 0
    best_score = 0

    for idx, row in enumerate(all_rows[:max_rows]):
        if not row:
            continue
        row_text = " ".join(c.lower() for c in row if c)
        score = sum(1 for kw in KNOWN_HEADERS if kw in row_text)
        if score >= 1 and score > best_score:
            best_score = score
            best_idx = idx

    return best_idx


def _clean_cell(cell) -> str:
    """Convert any openpyxl cell value to a clean string."""
    if cell is None:
        return ""
    return str(cell).strip()
